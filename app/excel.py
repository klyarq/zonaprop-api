import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = {
    "url":             "Link",
    "valor_x_m2":      "USD/m²",
    "price_value":     "Precio",
    "price_type":      "Moneda",
    "m2_ponderados":   "m² pond.",
    "m2_cubiertos":    "m² cub.",
    "m2_descubiertos": "m² desc.",
    "m2_totales":      "m² tot.",
    "ambientes":       "Amb.",
    "dormitorios":     "Dorm.",
    "banos":           "Baños",
    "cocheras":        "Coch.",
    "location":        "Ubicación",
    "expenses_value":  "Expensas",
    "expenses_type":   "Mon. exp.",
    "estado":          "Estado",
    "description":     "Descripción",
}

COL_WIDTHS = {
    "Link": 55, "USD/m²": 12, "Precio": 12, "Moneda": 9,
    "m² pond.": 10, "m² cub.": 10, "m² desc.": 10, "m² tot.": 10,
    "Amb.": 7, "Dorm.": 7, "Baños": 7, "Coch.": 7,
    "Ubicación": 32, "Expensas": 11, "Mon. exp.": 9,
    "Estado": 22, "Descripción": 65,
}

COLOR_HEADER_BG = "1F3864"
COLOR_NARANJA   = "FFC000"
COLOR_ROW_ALT   = "F2F2F2"
COLOR_BORDER    = "D0D0D0"


def _border():
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(records: list) -> bytes:
    df = pd.DataFrame(records)

    cols = [c for c in COLUMNS if c in df.columns]
    df = df[cols].rename(columns=COLUMNS)
    df = df.sort_values("USD/m²", ascending=True, na_position="last")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Propiedades")
        ws = writer.sheets["Propiedades"]

        header_fill  = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        naranja_fill = PatternFill("solid", fgColor=COLOR_NARANJA)
        alt_fill     = PatternFill("solid", fgColor=COLOR_ROW_ALT)
        no_fill      = PatternFill(fill_type=None)
        border       = _border()

        n_rows = df.shape[0]
        valor_idx = list(df.columns).index("USD/m²") + 1 if "USD/m²" in df.columns else None

        ws.row_dimensions[1].height = 28
        for ci, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(1, ci)
            is_v = ci == valor_idx
            cell.font      = Font(bold=True, size=11, color="000000" if is_v else "FFFFFF")
            cell.fill      = naranja_fill if is_v else header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        for ri in range(2, n_rows + 2):
            ws.row_dimensions[ri].height = 16
            alt = (ri % 2 == 0)
            for ci in range(1, len(df.columns) + 1):
                cell     = ws.cell(ri, ci)
                col_name = df.columns[ci - 1]
                is_v     = ci == valor_idx
                cell.border    = border
                cell.alignment = Alignment(
                    horizontal="left" if col_name in ("Link", "Ubicación", "Descripción", "Estado") else "center",
                    vertical="center",
                    wrap_text=(col_name == "Descripción"),
                )
                cell.fill = naranja_fill if is_v else (alt_fill if alt else no_fill)
                if is_v:
                    cell.font = Font(bold=True, size=11)
                if col_name in ("Precio", "Expensas", "USD/m²"):
                    cell.number_format = "#,##0"
                if col_name in ("m² pond.", "m² cub.", "m² desc.", "m² tot."):
                    cell.number_format = "#,##0.0"

        for ci, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col_name, 12)

        ws.freeze_panes = "A2"

    return buf.getvalue()
